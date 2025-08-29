import os
import time
import random
import re
from urllib.parse import urlparse
from difflib import SequenceMatcher

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from urllib.parse import urlparse

# Cargar API keys desde .env
try:
    load_dotenv()
    API_KEY = os.getenv("GOOGLE_API_KEY")
    CSE_ID = os.getenv("GOOGLE_CSE_ID")
    print("Environment variables loaded successfully")
except Exception as e:
    print(f"Warning: Error loading .env file: {e}")
    API_KEY = None
    CSE_ID = None

# -----------------------------
# Funciones de detección de duplicados
# -----------------------------
def similarity(a, b):
    """Calcula la similaridad entre dos strings"""
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()

def limpiar_nombre_empresa(nombre):
    """Limpia el nombre de la empresa para comparación"""
    if pd.isna(nombre):
        return ""
    
    # Convertir a string y limpiar
    nombre = str(nombre).lower().strip()
    
    # Remover sufijos corporativos comunes
    sufijos = [
        r'\b(inc|corp|corporation|ltd|limited|llc|llp|lp|co|company|enterprises|group|holding|international|global|worldwide|systems|solutions|software|technologies|technology|tech|services|consulting|digital|media|studios|games|entertainment|publishing|publishers|hardware|computers|computing)\b',
        r'\b(gmbh|ag|sa|srl|spa|bv|nv|oy|ab|as|\&|\+|\.|,)\b'
    ]
    
    for sufijo in sufijos:
        nombre = re.sub(sufijo, '', nombre)
    
    # Remover caracteres especiales y espacios extra
    nombre = re.sub(r'[^\w\s]', ' ', nombre)
    nombre = re.sub(r'\s+', ' ', nombre).strip()
    
    return nombre

def detectar_duplicados(df, name_col, threshold=0.85):
    """Detecta empresas duplicadas basándose en similaridad de nombres"""
    duplicados = []
    procesados = set()
    
    for i in range(len(df)):
        if i in procesados:
            continue
            
        nombre1 = limpiar_nombre_empresa(df.at[i, name_col])
        if not nombre1:
            continue
            
        grupo_duplicados = [i]
        
        for j in range(i + 1, len(df)):
            if j in procesados:
                continue
                
            nombre2 = limpiar_nombre_empresa(df.at[j, name_col])
            if not nombre2:
                continue
                
            if similarity(nombre1, nombre2) >= threshold:
                grupo_duplicados.append(j)
                procesados.add(j)
        
        if len(grupo_duplicados) > 1:
            duplicados.append(grupo_duplicados)
            for idx in grupo_duplicados:
                procesados.add(idx)
    
    return duplicados

# -----------------------------
# Funciones de categorización
# -----------------------------
def categorizar_empresa(nombre, website=""):
    """Categoriza una empresa basándose en su nombre y website"""
    if pd.isna(nombre):
        return "Unknown", "No data"
    
    nombre = str(nombre).lower()
    website = str(website).lower() if not pd.isna(website) else ""
    texto_completo = f"{nombre} {website}"
    
    # Palabras clave para categorización
    categorias = {
        # Publisher categories
        "Game Publisher": [
            "games", "gaming", "entertainment", "studios", "interactive", "digital entertainment",
            "game", "publisher", "publishing", "media", "activision", "electronic arts", "ubisoft"
        ],
        "Book Publisher": [
            "books", "publishing", "publications", "press", "editorial", "penguin", "harper",
            "macmillan", "scholastic", "textbook", "academic press"
        ],
        "Software Publisher": [
            "software", "applications", "apps", "programs", "development", "dev", "solutions",
            "microsoft", "adobe", "autodesk", "oracle"
        ],
        "Media Publisher": [
            "media", "news", "magazine", "newspaper", "broadcast", "streaming", "content",
            "netflix", "disney", "warner", "paramount"
        ],
        
        # Hardware categories
        "Computer Hardware": [
            "computers", "pc", "laptop", "desktop", "workstation", "server", "dell", "hp",
            "lenovo", "asus", "acer", "apple computer"
        ],
        "Components Provider": [
            "components", "parts", "processors", "cpu", "gpu", "memory", "storage", "motherboard",
            "intel", "amd", "nvidia", "corsair", "kingston", "seagate", "western digital"
        ],
        "Network Hardware": [
            "network", "networking", "router", "switch", "firewall", "wireless", "wifi",
            "cisco", "netgear", "tp-link", "ubiquiti", "juniper"
        ],
        "Mobile Hardware": [
            "mobile", "smartphone", "tablet", "phone", "cellular", "samsung", "apple iphone",
            "huawei", "xiaomi", "oneplus"
        ],
        
        # Service categories
        "Cloud Services": [
            "cloud", "hosting", "datacenter", "infrastructure", "saas", "paas", "iaas",
            "amazon aws", "google cloud", "microsoft azure", "digitalocean"
        ],
        "IT Services": [
            "consulting", "services", "integration", "support", "managed services",
            "ibm services", "accenture", "capgemini", "tcs"
        ],
        "Security Provider": [
            "security", "cybersecurity", "antivirus", "firewall", "encryption", "norton",
            "mcafee", "symantec", "kaspersky", "palo alto"
        ]
    }
    
    # Buscar coincidencias
    puntuaciones = {}
    for categoria, palabras in categorias.items():
        score = 0
        for palabra in palabras:
            if palabra in texto_completo:
                # Dar más peso si aparece en el nombre
                if palabra in nombre:
                    score += 3
                else:
                    score += 1
        puntuaciones[categoria] = score
    
    # Encontrar la mejor categoría
    if puntuaciones:
        mejor_categoria = max(puntuaciones.items(), key=lambda x: x[1])
        if mejor_categoria[1] > 0:
            # Generar descripción corta
            categoria = mejor_categoria[0]
            if "Publisher" in categoria:
                tipo = "Publisher"
                descripcion = categoria.replace(" Publisher", "").replace("_", " ")
            elif "Hardware" in categoria or "Provider" in categoria:
                tipo = "Hardware Provider"
                descripcion = categoria.replace(" Hardware", "").replace(" Provider", "").replace("_", " ")
            elif "Services" in categoria:
                tipo = "Service Provider"
                descripcion = categoria.replace(" Services", "").replace(" Provider", "").replace("_", " ")
            else:
                tipo = "Other"
                descripcion = categoria.replace("_", " ")
            
            return tipo, descripcion[:50]  # Limitar a 50 caracteres
    
    # Categoría por defecto
    return "Unknown", "Unclassified"

# -----------------------------
# Funciones de scoring para búsqueda
# -----------------------------
def generar_consultas_optimizadas(consulta_original: str):
    consulta_clean = consulta_original.strip()
    consultas = [
        f'"{consulta_clean}" official website',
        f'{consulta_clean} official site',
        f'{consulta_clean} homepage',
        f'{consulta_clean} company website',
        f'{consulta_clean} .com site:',
        f'{consulta_clean} software company',
        f'{consulta_clean} technology company'
    ]
    return consultas

def es_sitio_oficial(url: str, domain: str, title: str, snippet: str, consulta: str) -> int:
    score = 0
    consulta_lower = consulta.lower()
    domain_lower = domain.lower()
    title_lower = title.lower()
    snippet_lower = snippet.lower()
    
    consulta_base = re.sub(r'\b(software|hardware|inc|corp|ltd|llc|sa|srl|gmbh|ag)\b', '', consulta_lower).strip()
    palabras_consulta = [p for p in consulta_base.split() if len(p) > 2]

    # Palabra en dominio
    for palabra in palabras_consulta:
        if palabra in domain_lower:
            score += 25
    
    # Dominio exacto
    if any(domain_lower.startswith(f"{palabra}.") or f".{palabra}." in domain_lower for palabra in palabras_consulta):
        score += 40
    
    # TLD común
    if domain.endswith(('.com', '.net', '.org', '.io', '.tech', '.sl')):
        score += 15
    
    # Penalizar plataformas sociales
    social_platforms = [
        'facebook.com', 'twitter.com', 'linkedin.com', 'youtube.com', 'instagram.com',
        'wikipedia.org', 'crunchbase.com', 'bloomberg.com', 'reuters.com',
        'amazon.com', 'ebay.com', 'alibaba.com', 'github.com'
    ]
    for platform in social_platforms:
        if platform in domain_lower:
            score -= 30
            break
    
    # Palabras oficiales en título
    official_words = ['official', 'homepage', 'home page', 'corporate', 'company']
    if any(word in title_lower for word in official_words):
        score += 10
    
    # Palabras de consulta en título
    palabras_en_titulo = sum(1 for palabra in palabras_consulta if palabra in title_lower)
    score += palabras_en_titulo * 5
    
    # Penalizar subdominios
    subdomain_penalties = ['support.', 'help.', 'docs.', 'forum.', 'community.', 'blog.']
    if any(sub in domain_lower for sub in subdomain_penalties):
        score -= 10
    


# Penalizar URLs profundas excepto si contienen /#/
def calcular_score(url, score):
    parsed = urlparse(url)
    path = parsed.path

    if url.count('/') > 3 and '/#/' not in path:
        score -= 5

    if url.startswith('https://'):
        score += 5

    return max(0, min(100, score))


def seleccionar_mejor_url_oficial(consulta: str, candidatos):
    if not candidatos:
        return None, "no candidates"
    
    scored_candidates = []
    for item in candidatos:
        url = item.get("href", "")
        title = item.get("title", "")
        snippet = item.get("snippet", "")
        domain = item.get("displayLink", "")
        
        if not url:
            continue
        
        score = es_sitio_oficial(url, domain, title, snippet, consulta)
        scored_candidates.append({"score": score, "url": url, "domain": domain})
    
    if not scored_candidates:
        return None, "no valid candidates"
    
    scored_candidates.sort(key=lambda x: x['score'], reverse=True)
    best = scored_candidates[0]
    return best['url'], f"score {best['score']}, domain: {best['domain']}"

# -----------------------------
# Función para Google CSE
# -----------------------------
def buscar_con_google_cse_multiples(consultas):
    # Verificar que las API keys estén disponibles
    if not API_KEY or not CSE_ID:
        print("❌ Google API keys not found. Skipping search.")
        return []
    
    todos_candidatos = []
    urls_vistas = set()
    
    for query in consultas[:3]:  # hasta 3 consultas
        try:
            url = "https://www.googleapis.com/customsearch/v1"
            params = {
                'key': API_KEY,
                'cx': CSE_ID,
                'q': query,
                'num': 5,
                'safe': 'medium'
            }
            
            response = requests.get(url, params=params, timeout=15)
            if response.status_code == 200:
                data = response.json()
                for item in data.get('items', []):
                    href = item.get("link", "")
                    if href and href not in urls_vistas:
                        candidato = {
                            "title": item.get("title", ""),
                            "href": href,
                            "snippet": item.get("snippet", ""),
                            "displayLink": item.get("displayLink", "")
                        }
                        todos_candidatos.append(candidato)
                        urls_vistas.add(href)
            elif response.status_code == 429:
                print("Rate limit, esperando 30s...")
                time.sleep(30)
            
            time.sleep(1)  # delay entre consultas
        except Exception as e:
            print(f"Error en consulta '{query}': {e}")
    
    return todos_candidatos

# -----------------------------
# Función para verificar URLs (modificada para True/False)
# -----------------------------
def verificar_url(url):
    """Verifica si una URL es accesible y funciona correctamente"""
    if pd.isna(url) or not str(url).strip():
        return False, "Empty URL"
    
    url_str = str(url).strip()
    
    # Agregar http si no tiene protocolo
    if not url_str.startswith(('http://', 'https://')):
        url_str = 'http://' + url_str
    
    try:
        response = requests.get(url_str, timeout=10, allow_redirects=True)
        if response.status_code == 404:
            return False, "Error 404"
        elif response.status_code >= 400:
            return False, f"Error {response.status_code}"
        else:
            return True, "OK"
    except requests.exceptions.Timeout:
        return False, "Timeout"
    except requests.exceptions.ConnectionError:
        return False, "Connection Error"
    except requests.exceptions.RequestException as e:
        return False, f"Request Error: {str(e)[:50]}"
    except Exception as e:
        return False, f"General Error: {str(e)[:50]}"

# -----------------------------
# Función principal
# -----------------------------
def main():
    # Configuración de archivos
    input_file = "./app/publishers.csv"  # Archivo CSV de entrada
    output_excel = "./app/publishers_verified.xlsx"
    
    # Verificar que el archivo de entrada existe
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        print("Please make sure the CSV file exists in the correct location.")
        return
    
    # Cargar datos desde CSV con manejo de encoding
    print("Cargando archivo CSV...")
    
    # Intentar diferentes encodings
    encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-8-sig']
    df = None
    
    for encoding in encodings_to_try:
        try:
            print(f"Trying encoding: {encoding}")
            df = pd.read_csv(input_file, encoding=encoding)
            print(f"✅ Successfully loaded CSV with encoding: {encoding}")
            break
        except UnicodeDecodeError as e:
            print(f"❌ Failed with {encoding}: {e}")
            continue
        except Exception as e:
            print(f"❌ Error with {encoding}: {e}")
            continue
    
    if df is None:
        print("❌ Could not load CSV file with any encoding. Please check the file.")
        return
    
    # Buscar columnas de nombre y website
    name_col = None
    website_col = None
    
    for col in df.columns:
        col_lower = col.lower()
        if any(word in col_lower for word in ['company', 'name', 'publisher', 'empresa', 'nombre']):
            if name_col is None:
                name_col = col
        elif any(word in col_lower for word in ['website', 'url', 'site', 'web', 'sitio']):
            if website_col is None:
                website_col = col
    
    if name_col is None:
        name_col = df.columns[0]  # usar primera columna como fallback
        print(f"Using '{name_col}' as name column")
    
    if website_col is None:
        website_col = 'Website'  # crear nueva columna
        df[website_col] = None
        print(f"Creating new column '{website_col}'")
    
    print(f"Name column: {name_col}")
    print(f"Website column: {website_col}")
    
    # FASE 0: Detectar duplicados
    print("\n" + "="*50)
    print("PHASE 0: Detecting duplicates...")
    print("="*50)
    
    duplicados = detectar_duplicados(df, name_col)
    df['is_duplicate'] = False
    df['duplicate_group'] = None
    
    if duplicados:
        print(f"Found {len(duplicados)} groups of duplicates:")
        for i, grupo in enumerate(duplicados):
            print(f"\nGroup {i+1}:")
            for idx in grupo:
                nombre = df.at[idx, name_col]
                website = df.at[idx, website_col] if not pd.isna(df.at[idx, website_col]) else "No URL"
                print(f"  - Row {idx+1}: {nombre} ({website})")
                df.at[idx, 'is_duplicate'] = True
                df.at[idx, 'duplicate_group'] = f"Group_{i+1}"
    else:
        print("No duplicates detected")
    
    # Agregar columnas de resultado
    df['found_url'] = None
    df['search_notes'] = None
    df['url_works'] = None  # Será True/False
    df['verification_status'] = None
    df['company_type'] = None
    df['category_description'] = None
    
    # FASE 1: Buscar URLs faltantes
    print("\n" + "="*50)
    print("PHASE 1: Searching for missing URLs...")
    print("="*50)
    
    filas_sin_url = df[df[website_col].isna() | (df[website_col].str.strip() == '')].index
    print(f"Found {len(filas_sin_url)} rows without URL")
    
    for idx in filas_sin_url:
        if idx >= len(df):
            continue
            
        consulta = str(df.at[idx, name_col]).strip()
        if not consulta or consulta.lower() == 'nan':
            df.at[idx, 'search_notes'] = "empty name"
            continue

        print(f"\nSearching official site for: {consulta}")
        consultas = generar_consultas_optimizadas(consulta)
        candidatos = buscar_con_google_cse_multiples(consultas)
        url, notas = seleccionar_mejor_url_oficial(consulta, candidatos)
        
        if url:
            df.at[idx, website_col] = url
            df.at[idx, 'found_url'] = url
        
        df.at[idx, 'search_notes'] = notas
        print(f"→ {url} ({notas})")
        
        # Delay entre búsquedas para evitar rate limits
        time.sleep(random.uniform(2, 4))
    
    # FASE 2: Verificar todas las URLs
    print("\n" + "="*50)
    print("PHASE 2: Verifying URL functionality...")
    print("="*50)
    
    total_urls = len(df[df[website_col].notna() & (df[website_col].str.strip() != '')])
    print(f"Verifying {total_urls} URLs...")
    
    urls_con_error = []
    
    for idx, row in df.iterrows():
        url = row[website_col]
        if pd.isna(url) or not str(url).strip():
            df.at[idx, 'url_works'] = "False"
            df.at[idx, 'verification_status'] = "No URL"
            continue
        
        print(f"Verifying: {url}")
        funciona, estado = verificar_url(url)
        
        df.at[idx, 'url_works'] = "True" if funciona else "False"  # Force English text
        df.at[idx, 'verification_status'] = estado
        
        if not funciona:
            urls_con_error.append(idx)
            print(f"❌ {url} - {estado}")
        else:
            print(f"✅ {url} - OK")
        
        time.sleep(0.5)  # pequeño delay
    
    # FASE 3: Categorizar empresas
    print("\n" + "="*50)
    print("PHASE 3: Categorizing companies...")
    print("="*50)
    
    for idx, row in df.iterrows():
        nombre = row[name_col]
        website = row[website_col]
        
        tipo, descripcion = categorizar_empresa(nombre, website)
        df.at[idx, 'company_type'] = tipo
        df.at[idx, 'category_description'] = descripcion
        
        print(f"{nombre} → {tipo}: {descripcion}")
    
    # FASE 4: Guardar y resaltar errores
    print("\n" + "="*50)
    print("PHASE 4: Saving results and highlighting errors...")
    print("="*50)
    
    # Guardar a Excel
    df.to_excel(output_excel, index=False)
    
    # Aplicar formato de color
    if urls_con_error or duplicados:
        wb = load_workbook(output_excel)
        ws = wb.active
        
        # Definir colores
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo para errores
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # Verde claro para encontradas
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")    # Azul claro para duplicados
        
        print(f"Highlighting {len(urls_con_error)} rows with errors...")
        
        for row_idx in range(2, ws.max_row + 1):  # empezar desde fila 2 (después del header)
            df_row_idx = row_idx - 2  # ajustar índice para DataFrame
            
            if df_row_idx < len(df):
                # Prioridad: duplicados (azul) > errores (amarillo) > encontradas (verde)
                if df.at[df_row_idx, 'is_duplicate']:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = blue_fill
                elif not df.at[df_row_idx, 'url_works']:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = yellow_fill
                elif pd.notna(df.at[df_row_idx, 'found_url']):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = green_fill
        
        wb.save(output_excel)
        print(f"Format applied successfully")
    
    # RESUMEN FINAL
    print("\n" + "="*60)
    print("FINAL SUMMARY")
    print("="*60)
    
    total_registros = len(df)
    duplicados_totales = len(df[df['is_duplicate'] == True])
    urls_encontradas = len(df[df['found_url'].notna()])
    urls_funcionando = len(df[df['url_works'] == True])
    urls_con_errores = len(df[df['url_works'] == False])
    urls_sin_verificar = total_registros - urls_funcionando - urls_con_errores
    
    print(f"📊 Total records: {total_registros}")
    print(f"🔄 Duplicates detected: {duplicados_totales}")
    print(f"🔍 Found URLs (new): {urls_encontradas}")
    print(f"✅ Working URLs: {urls_funcionando}")
    print(f"❌ URLs with errors: {urls_con_errores}")
    print(f"⚪ No URL to verify: {urls_sin_verificar}")
    
    # Resumen por categorías
    print(f"\n📂 Company categorization:")
    categorias = df['company_type'].value_counts()
    for categoria, count in categorias.items():
        print(f"   - {categoria}: {count}")
    
    print(f"\n📁 File saved: '{output_excel}'")
    print("🎨 Format applied:")
    print("   - Blue: Duplicate companies")
    print("   - Yellow: URLs with errors")
    print("   - Light green: New URLs found")
    print("   - No color: Existing URLs that work")

if __name__ == "__main__":
    main()