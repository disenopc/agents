import os
import time
import random
import re
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz, process

# -----------------------------
# Cargar API keys desde .env
# -----------------------------
try:
    load_dotenv()
    API_KEY = os.getenv("GOOGLE_API_KEY")
    CSE_ID = os.getenv("GOOGLE_CSE_ID")
    print("Environment variables loaded successfully")
except Exception as e:
    print(f"Warning: Error loading .env file: {e}")
    API_KEY = None
    CSE_ID = None

# -----------------------------
# Funciones de detección de duplicados
# -----------------------------
def limpiar_nombre_empresa(nombre):
    """Limpia el nombre de la empresa para comparación"""
    if pd.isna(nombre):
        return ""
    
    nombre = str(nombre).lower().strip()
    sufijos = [
        r'\b(inc|corp|corporation|ltd|limited|llc|llp|lp|co|company|enterprises|group|holding|international|global|worldwide|systems|solutions|software|technologies|technology|tech|services|consulting|digital|media|studios|games|entertainment|publishing|publishers|hardware|computers|computing)\b',
        r'\b(gmbh|ag|sa|srl|spa|bv|nv|oy|ab|as|\&|\+|\.|,)\b'
    ]
    for sufijo in sufijos:
        nombre = re.sub(sufijo, '', nombre)
    nombre = re.sub(r'[^\w\s]', ' ', nombre)
    nombre = re.sub(r'\s+', ' ', nombre).strip()
    return nombre

def detectar_duplicados(df, name_col, threshold=85):
    """Detecta duplicados usando rapidfuzz (más rápido que difflib)"""
    df['clean_name'] = df[name_col].apply(limpiar_nombre_empresa)
    duplicados = []
    seen = set()

    for i, name in enumerate(df['clean_name']):
        if i in seen or not name:
            continue
        matches = process.extract(name, df['clean_name'], scorer=fuzz.ratio, limit=None)
        grupo = [j for j, score in [(idx, s) for idx, s in enumerate(dict(matches).values())] if score >= threshold and j != i]
        if grupo:
            duplicados.append([i] + grupo)
            seen.update(grupo)
    return duplicados

# -----------------------------
# Funciones de categorización
# -----------------------------
def categorizar_empresa(nombre, website=""):
    if pd.isna(nombre):
        return "Unknown", "No data"
    
    nombre = str(nombre).lower()
    website = str(website).lower() if not pd.isna(website) else ""
    texto_completo = f"{nombre} {website}"
    
    categorias = {
        "Game Publisher": ["games", "gaming", "entertainment", "studios", "interactive", "digital entertainment",
                           "game", "publisher", "publishing", "media", "activision", "electronic arts", "ubisoft"],
        "Book Publisher": ["books", "publishing", "publications", "press", "editorial", "penguin", "harper",
                           "macmillan", "scholastic", "textbook", "academic press"],
        "Software Publisher": ["software", "applications", "apps", "programs", "development", "dev", "solutions",
                               "microsoft", "adobe", "autodesk", "oracle"],
        "Media Publisher": ["media", "news", "magazine", "newspaper", "broadcast", "streaming", "content",
                            "netflix", "disney", "warner", "paramount"],
        "Computer Hardware": ["computers", "pc", "laptop", "desktop", "workstation", "server", "dell", "hp",
                              "lenovo", "asus", "acer", "apple computer"],
        "Components Provider": ["components", "parts", "processors", "cpu", "gpu", "memory", "storage", "motherboard",
                                "intel", "amd", "nvidia", "corsair", "kingston", "seagate", "western digital"],
        "Network Hardware": ["network", "networking", "router", "switch", "firewall", "wireless", "wifi",
                             "cisco", "netgear", "tp-link", "ubiquiti", "juniper"],
        "Mobile Hardware": ["mobile", "smartphone", "tablet", "phone", "cellular", "samsung", "apple iphone",
                            "huawei", "xiaomi", "oneplus"],
        "Cloud Services": ["cloud", "hosting", "datacenter", "infrastructure", "saas", "paas", "iaas",
                           "amazon aws", "google cloud", "microsoft azure", "digitalocean"],
        "IT Services": ["consulting", "services", "integration", "support", "managed services",
                        "ibm services", "accenture", "capgemini", "tcs"],
        "Security Provider": ["security", "cybersecurity", "antivirus", "firewall", "encryption", "norton",
                              "mcafee", "symantec", "kaspersky", "palo alto"]
    }
    
    puntuaciones = {}
    for categoria, palabras in categorias.items():
        score = 0
        for palabra in palabras:
            if palabra in texto_completo:
                score += 3 if palabra in nombre else 1
        puntuaciones[categoria] = score
    
    if puntuaciones:
        mejor_categoria = max(puntuaciones.items(), key=lambda x: x[1])
        if mejor_categoria[1] > 0:
            categoria = mejor_categoria[0]
            if "Publisher" in categoria:
                return "Publisher", categoria.replace(" Publisher", "")
            elif "Hardware" in categoria or "Provider" in categoria:
                return "Hardware Provider", categoria.replace(" Hardware", "").replace(" Provider", "")
            elif "Services" in categoria:
                return "Service Provider", categoria.replace(" Services", "").replace(" Provider", "")
            else:
                return "Other", categoria
    return "Unknown", "Unclassified"

# -----------------------------
# Funciones de scoring / búsqueda
# -----------------------------
def generar_consultas_optimizadas(consulta_original: str):
    consulta_clean = consulta_original.strip()
    return [
        f'"{consulta_clean}" official website',
        f'{consulta_clean} official site',
        f'{consulta_clean} homepage',
        f'{consulta_clean} company website',
        f'{consulta_clean} software company',
        f'{consulta_clean} technology company'
    ]

def es_sitio_oficial(url: str, domain: str, title: str, snippet: str, consulta: str) -> int:
    score = 0
    consulta_lower = consulta.lower()
    domain_lower = domain.lower()
    title_lower = title.lower()
    snippet_lower = snippet.lower()
    
    consulta_base = re.sub(r'\b(software|hardware|inc|corp|ltd|llc|sa|srl|gmbh|ag)\b', '', consulta_lower).strip()
    palabras_consulta = [p for p in consulta_base.split() if len(p) > 2]

    for palabra in palabras_consulta:
        if palabra in domain_lower:
            score += 25
    if any(domain_lower.startswith(f"{palabra}.") or f".{palabra}." in domain_lower for palabra in palabras_consulta):
        score += 40
    if domain.endswith(('.com', '.net', '.org', '.io', '.tech')):
        score += 15
    social_platforms = ['facebook.com','twitter.com','linkedin.com','youtube.com','instagram.com','wikipedia.org',
                        'crunchbase.com','bloomberg.com','reuters.com','amazon.com','ebay.com','alibaba.com','github.com']
    if any(platform in domain_lower for platform in social_platforms):
        score -= 30
    if any(word in title_lower for word in ['official', 'homepage', 'corporate', 'company']):
        score += 10
    score += sum(1 for palabra in palabras_consulta if palabra in title_lower) * 5
    return max(0, min(100, score))

def seleccionar_mejor_url_oficial(consulta: str, candidatos):
    if not candidatos:
        return None, "no candidates"
    scored_candidates = []
    for item in candidatos:
        url = item.get("href", "")
        if not url:
            continue
        score = es_sitio_oficial(url, item.get("displayLink",""), item.get("title",""), item.get("snippet",""), consulta)
        scored_candidates.append({"score": score, "url": url, "domain": item.get("displayLink","")})
    if not scored_candidates:
        return None, "no valid candidates"
    best = max(scored_candidates, key=lambda x: x['score'])
    return best['url'], f"score {best['score']}, domain: {best['domain']}"

def buscar_con_google_cse_multiples(consultas):
    if not API_KEY or not CSE_ID:
        return []
    todos_candidatos, urls_vistas = [], set()
    for query in consultas[:2]:  # menos consultas
        try:
            url = "https://www.googleapis.com/customsearch/v1"
            params = {'key': API_KEY, 'cx': CSE_ID, 'q': query, 'num': 5, 'safe': 'medium'}
            response = requests.get(url, params=params, timeout=15)
            if response.status_code == 200:
                data = response.json()
                for item in data.get('items', []):
                    href = item.get("link", "")
                    if href and href not in urls_vistas:
                        todos_candidatos.append({
                            "title": item.get("title", ""),
                            "href": href,
                            "snippet": item.get("snippet", ""),
                            "displayLink": item.get("displayLink", "")
                        })
                        urls_vistas.add(href)
            time.sleep(0.5)
        except Exception as e:
            print(f"Error en consulta '{query}': {e}")
    return todos_candidatos

# -----------------------------
# Funciones de verificación
# -----------------------------
def verificar_url(url):
    if pd.isna(url) or not str(url).strip():
        return False, "Empty URL"
    url_str = str(url).strip()
    if not url_str.startswith(('http://','https://')):
        url_str = 'http://' + url_str
    try:
        response = requests.get(url_str, timeout=10, allow_redirects=True)
        if response.status_code >= 400:
            return False, f"Error {response.status_code}"
        return True, "OK"
    except requests.exceptions.Timeout:
        return False, "Timeout"
    except requests.exceptions.ConnectionError:
        return False, "Connection Error"
    except Exception as e:
        return False, f"Error: {str(e)[:50]}"

def verificar_urls_batch(df, website_col):
    resultados = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_idx = {
            executor.submit(verificar_url, df.at[idx, website_col]): idx
            for idx in df.index
            if pd.notna(df.at[idx, website_col]) and str(df.at[idx, website_col]).strip()
        }
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            try:
                funciona, estado = future.result()
                resultados[idx] = (funciona, estado)
            except Exception as e:
                resultados[idx] = (False, f"Error: {e}")
    for idx, (funciona, estado) in resultados.items():
        df.at[idx, 'url_works'] = "True" if funciona else "False"
        df.at[idx, 'verification_status'] = estado
    return df

# -----------------------------
# Función principal
# -----------------------------
def main():
    input_file = "./app/publishers.csv"
    output_excel = "./app/publishers_verified.xlsx"
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        return
    
    print("Cargando archivo CSV...")
    encodings_to_try = ['utf-8','latin-1','cp1252','iso-8859-1','utf-8-sig']
    df = None
    for encoding in encodings_to_try:
        try:
            df = pd.read_csv(input_file, encoding=encoding)
            print(f"✅ Loaded CSV with encoding: {encoding}")
            break
        except Exception as e:
            continue
    if df is None:
        print("❌ Could not load CSV file")
        return
    
    name_col, website_col = None, None
    for col in df.columns:
        col_lower = col.lower()
        if any(word in col_lower for word in ['company','name','publisher','empresa','nombre']):
            if name_col is None:
                name_col = col
        elif any(word in col_lower for word in ['website','url','site','web','sitio']):
            if website_col is None:
                website_col = col
    if name_col is None:
        name_col = df.columns[0]
    if website_col is None:
        website_col = 'Website'
        df[website_col] = None
    
    print("Detectando duplicados...")
    duplicados = detectar_duplicados(df, name_col)
    df['is_duplicate'] = False
    df['duplicate_group'] = None
    for i, grupo in enumerate(duplicados):
        for idx in grupo:
            df.at[idx, 'is_duplicate'] = True
            df.at[idx, 'duplicate_group'] = f"Group_{i+1}"
    
    df['found_url'], df['search_notes'], df['url_works'], df['verification_status'] = None, None, None, None
    df['company_type'], df['category_description'] = None, None
    
    print("Buscando URLs faltantes...")
    filas_sin_url = df[df[website_col].isna() | (df[website_col].str.strip() == '')].index
    cache_busquedas = {}
    for idx in filas_sin_url:
        consulta = str(df.at[idx, name_col]).strip()
        if not consulta or consulta.lower() == 'nan':
            df.at[idx, 'search_notes'] = "empty name"
            continue
        if consulta in cache_busquedas:
            url, notas = cache_busquedas[consulta]
        else:
            consultas = generar_consultas_optimizadas(consulta)
            candidatos = buscar_con_google_cse_multiples(consultas)
            url, notas = seleccionar_mejor_url_oficial(consulta, candidatos)
            cache_busquedas[consulta] = (url, notas)
        if url:
            df.at[idx, website_col] = url
            df.at[idx, 'found_url'] = url
        df.at[idx, 'search_notes'] = notas
    
    print("Verificando URLs en paralelo...")
    df = verificar_urls_batch(df, website_col)
    
    print("Categorizar empresas...")
    df[['company_type','category_description']] = df.apply(
        lambda row: pd.Series(categorizar_empresa(row[name_col], row[website_col])),
        axis=1
    )
    
    print("Guardando Excel...")
    df.to_excel(output_excel, index=False)
    
    wb = load_workbook(output_excel)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    for row_idx in range(2, ws.max_row+1):
        df_row_idx = row_idx - 2
        if df_row_idx < len(df):
            if df.at[df_row_idx, 'is_duplicate']:
                for col in range(1, ws.max_column+1):
                    ws.cell(row=row_idx, column=col).fill = blue_fill
            elif df.at[df_row_idx, 'url_works'] == "False":
                for col in range(1, ws.max_column+1):
                    ws.cell(row=row_idx, column=col).fill = yellow_fill
            elif pd.notna(df.at[df_row_idx, 'found_url']):
                for col in range(1, ws.max_column+1):
                    ws.cell(row=row_idx, column=col).fill = green_fill
    
    wb.save(output_excel)
    print("✅ Archivo guardado y coloreado")

if __name__ == "__main__":
    main()
