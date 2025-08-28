import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Cargar archivo CSV con columna 'website'
df = pd.read_csv('./app/publishers.csv')

# Crear una columna para marcar errores
df['error_404'] = False

# Verificar cada URL
for i, url in enumerate(df['Website']):
    if not url.startswith(('http://', 'https://')):
        url = 'http://' + url  # agrega http si no tiene
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 404:
            df.at[i, 'error_404'] = True
            print(f"URL 404 encontrada: {url}")
    except requests.RequestException as e:
        df.at[i, 'error_404'] = True
        print(f"Error al acceder a {url}: {e}")


# Guardar a Excel
excel_file = 'clientes_verificados.xlsx'
df.to_excel(excel_file, index=False)

# Abrir el Excel para aplicar color
wb = load_workbook(excel_file)
ws = wb.active

# Definir color amarillo
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Obtener la columna exacta de 'error_404'
col_index = df.columns.get_loc('error_404') + 1  # openpyxl empieza en 1

# Aplicar color a filas con error 404
for row in range(2, ws.max_row + 1):  # asumiendo que la primera fila es encabezado
    if ws.cell(row=row, column=col_index).value == True:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = yellow_fill

wb.save(excel_file)
print("Archivo clientes_verificados.xlsx generado con errores resaltados.")
